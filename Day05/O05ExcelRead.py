
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

print(ws.dimensions)

datarange = ws[ws.dimensions]
# print(f"datarange :{datarange}")

for c1, c2, c3, c4, c5 in datarange:
    print("{0:5}\t\t{1:5}\t\t{2:5}\t\t{3:5}\t\t{4:5}".format(c1.value, c2.value,
                                                        c3.value, c4.value, c5.value))


