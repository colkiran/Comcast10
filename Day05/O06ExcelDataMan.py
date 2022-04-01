
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

ws = wb.active

print(ws.dimensions)

for row in ws.iter_rows(min_row=6, min_col=1, max_row=11, max_col=1):
    for cell in row:
        # print(cell.value)
        if cell.value == "Lara":
            cell.value = "brain lara".upper()

wb.save("FirstExcel.xlsx")