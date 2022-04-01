
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Players")

wb.active = wb["Players"]

ws = wb.active

cell = ws["A5"]

data = [
    ['Playername', 'Age', 'Runs', 'Matches', 'Country'],
    ['Sachin', 49, 38950, 975, 'India'],
    ['Lara', 51, 25800, 730, 'West Indies'],
    ['Ponting', 47, 28670, 680, 'Australia'],
    ['Jayasurya', 50, 24500, 584, 'Sri Lanka'],
    ['Inzamam', 46, 27800, 759, 'Pakistan']
]

print(data)

for row in data:
    ws.append(row)

wb.save("FirstExcel.xlsx")
