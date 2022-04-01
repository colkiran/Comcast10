"""
Excel
-----
file    -   Workbook
pages   -   Worksheet
cell    -   intersection of row and column
            cell will have and address, which is used to access the value in the cell
            its a combination of colname + rowno => A + 10 => A10

row     -   numbered - 1..1048576
cols    -   named   -  A..Z, AA..AZ,...XFD (16384)

"""

from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "comcast"

ws['C5'] = "Hello World"

ws['C6'] = 12845

ws["C7"] = datetime.now()

wb.save("FirstExcel.xlsx")
